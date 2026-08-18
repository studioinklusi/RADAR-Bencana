# -*- coding: utf-8 -*-
"""
QGIS Processing Algorithm untuk menghitung estimasi jumlah penduduk terdampak bencana 
berdasarkan metode dasimetrik raster dan mengekspor hasilnya ke format Excel & HTML Interaktif.
"""

import os
from qgis.PyQt.QtCore import QCoreApplication
from qgis.core import (QgsProcessing,
                       QgsProcessingAlgorithm,
                       QgsProcessingParameterRasterLayer,
                       QgsProcessingParameterFeatureSource,
                       QgsProcessingParameterField,
                       QgsProcessingParameterFolderDestination,
                       QgsProcessingParameterFileDestination,
                       QgsProcessingParameterNumber,
                       QgsProcessingException,
                       QgsCoordinateTransform,
                       QgsCoordinateReferenceSystem,
                       QgsProject,
                       QgsFeatureRequest)
import processing

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DasimetrikBencanaAlgorithm(QgsProcessingAlgorithm):
    P_RASTER_BENCANA = 'P_RASTER_BENCANA'
    P_RASTER_PENDUDUK = 'P_RASTER_PENDUDUK'
    P_VEKTOR_DESA = 'P_VEKTOR_DESA'
    P_FIELD_ID = 'P_FIELD_ID'
    P_FIELD_DESA = 'P_FIELD_DESA'
    P_FIELD_KEC = 'P_FIELD_KEC'
    P_DECIMAL_LUAS = 'P_DECIMAL_LUAS'
    P_OUT_FOLDER = 'P_OUT_FOLDER'
    P_OUT_EXCEL = 'P_OUT_EXCEL'
    P_OUT_HTML = 'P_OUT_HTML' # Parameter baru untuk HTML

    def tr(self, string):
        return QCoreApplication.translate('Processing', string)

    def createInstance(self):
        return DasimetrikBencanaAlgorithm()

    def name(self):
        return 'dasimetrikbencana'

    def displayName(self):
        return self.tr('Hitung Dampak Bencana Dasimetrik (Excel & HTML)')

    def group(self):
        return self.tr('Analisis Kebencanaan')

    def groupId(self):
        return 'analisiskebencanaan'

    def shortHelpString(self):
        return self.tr("Algoritma ini menghitung luasan kelas bahaya dan estimasi jumlah "
                       "penduduk terpapar per desa, lalu mengekspornya ke Excel dan HTML Interaktif.\n\n"
                       "Catatan: File HTML interaktif membutuhkan koneksi internet saat dibuka "
                       "untuk memuat pustaka tampilan (Bootstrap & DataTables).")

    def initAlgorithm(self, config=None):
        self.addParameter(QgsProcessingParameterRasterLayer(
            self.P_RASTER_BENCANA,
            self.tr('Raster Indeks Bencana')
        ))
        self.addParameter(QgsProcessingParameterRasterLayer(
            self.P_RASTER_PENDUDUK,
            self.tr('Raster Dasimetrik Penduduk')
        ))
        
        self.addParameter(QgsProcessingParameterFeatureSource(
            self.P_VEKTOR_DESA,
            self.tr('SHP Administrasi Desa (Polygon)'),
            [QgsProcessing.TypeVectorPolygon]
        ))

        self.addParameter(QgsProcessingParameterField(
            self.P_FIELD_ID,
            self.tr('Field ID Unik Desa (misal Kode BPS - Numerik/Integer)'),
            parentLayerParameterName=self.P_VEKTOR_DESA
        ))
        self.addParameter(QgsProcessingParameterField(
            self.P_FIELD_DESA,
            self.tr('Field Nama Desa (Untuk Laporan)'),
            parentLayerParameterName=self.P_VEKTOR_DESA
        ))
        self.addParameter(QgsProcessingParameterField(
            self.P_FIELD_KEC,
            self.tr('Field Nama Kecamatan (Untuk Laporan)'),
            parentLayerParameterName=self.P_VEKTOR_DESA
        ))

        self.addParameter(QgsProcessingParameterNumber(
            self.P_DECIMAL_LUAS,
            self.tr('Jumlah Desimal Pembulatan Luas (Ha)'),
            type=QgsProcessingParameterNumber.Integer,
            defaultValue=2,
            minValue=0
        ))

        self.addParameter(QgsProcessingParameterFolderDestination(
            self.P_OUT_FOLDER,
            self.tr('Folder Output Sementara (Untuk Raster Antara)')
        ))
        
        self.addParameter(QgsProcessingParameterFileDestination(
            self.P_OUT_EXCEL,
            self.tr('File Output Laporan (Excel .xlsx)'),
            fileFilter='Excel files (*.xlsx)'
        ))

        # Tambahan Parameter HTML
        self.addParameter(QgsProcessingParameterFileDestination(
            self.P_OUT_HTML,
            self.tr('File Output Laporan Interaktif (HTML .html) [Opsional]'),
            fileFilter='HTML files (*.html)',
            optional=True
        ))

    def processAlgorithm(self, parameters, context, feedback):
        if not HAS_OPENPYXL:
            raise QgsProcessingException("Library 'openpyxl' tidak ditemukan. Silakan install "
                                         "melalui OSGeo4W Shell.")

        raster_bencana = self.parameterAsRasterLayer(parameters, self.P_RASTER_BENCANA, context)
        raster_penduduk = self.parameterAsRasterLayer(parameters, self.P_RASTER_PENDUDUK, context)
        vektor_desa = self.parameterAsSource(parameters, self.P_VEKTOR_DESA, context)
        
        field_id = self.parameterAsString(parameters, self.P_FIELD_ID, context)
        field_desa = self.parameterAsString(parameters, self.P_FIELD_DESA, context)
        field_kec = self.parameterAsString(parameters, self.P_FIELD_KEC, context)
        
        desimal_luas = self.parameterAsInt(parameters, self.P_DECIMAL_LUAS, context)
        out_folder = self.parameterAsString(parameters, self.P_OUT_FOLDER, context)
        out_excel = self.parameterAsString(parameters, self.P_OUT_EXCEL, context)
        out_html = self.parameterAsString(parameters, self.P_OUT_HTML, context)

        if not raster_bencana or not raster_penduduk or not vektor_desa:
            raise QgsProcessingException("Input layer tidak valid atau kosong.")

        # --- VALIDASI CRS ---
        crs_bencana = raster_bencana.crs()
        # isProjected() tidak tersedia di QGIS 3.38+. Gunakan kebalikan isGeographic()
        # isGeographic() = True → CRS masih lat/lon (belum projected) → ditolak
        if not crs_bencana.isValid() or crs_bencana.isGeographic():
            raise QgsProcessingException(f"CRS Raster Indeks Bencana ({crs_bencana.authid()}) tidak valid atau bukan projected.")
        
        pixel_size_x = raster_bencana.rasterUnitsPerPixelX()
        pixel_size_y = raster_bencana.rasterUnitsPerPixelY()
        extent_bencana = raster_bencana.extent()
        
        if feedback.isCanceled(): return {}
        
        # Alignment CRS Layer Penduduk
        crs_penduduk = raster_penduduk.crs()
        aligned_penduduk_path = os.path.join(out_folder, "penduduk_aligned.tif")
        if crs_penduduk != crs_bencana or raster_penduduk.extent() != extent_bencana or \
           raster_penduduk.rasterUnitsPerPixelX() != pixel_size_x:
            feedback.pushInfo("Menyelaraskan Raster Penduduk...")
            # FIX #2: Gunakan Nearest Neighbour (RESAMPLING=0) untuk raster penduduk
            # karena data populasi bersifat diskrit/count. Bilinear interpolation
            # dapat mengubah total jumlah penduduk secara tidak akurat.
            warp_params = {
                'INPUT': parameters[self.P_RASTER_PENDUDUK],
                'TARGET_CRS': crs_bencana.authid(),
                'RESAMPLING': 0,  # 0 = Nearest Neighbour (sebelumnya: 1 = Bilinear)
                'NODATA': 0,
                'TARGET_RESOLUTION': pixel_size_x,
                'TARGET_EXTENT': f"{extent_bencana.xMinimum()},{extent_bencana.xMaximum()},{extent_bencana.yMinimum()},{extent_bencana.yMaximum()}",
                'OUTPUT': aligned_penduduk_path
            }
            res_warp = processing.run("gdal:warpreproject", warp_params, context=context, feedback=feedback, is_child_algorithm=True)
            aligned_penduduk = res_warp['OUTPUT']
        else:
            aligned_penduduk = parameters[self.P_RASTER_PENDUDUK]

        if feedback.isCanceled(): return {}

        # Alignment CRS Vektor Desa
        if vektor_desa.sourceCrs() != crs_bencana:
            feedback.pushInfo("Reproject Vektor Desa...")
            rep_params = {
                'INPUT': parameters[self.P_VEKTOR_DESA],
                'TARGET_CRS': crs_bencana.authid(),
                'OUTPUT': 'memory:'
            }
            res_rep = processing.run("native:reprojectlayer", rep_params, context=context, feedback=feedback, is_child_algorithm=True)
            aligned_vektor = res_rep['OUTPUT']
        else:
            aligned_vektor = parameters[self.P_VEKTOR_DESA]

        if feedback.isCanceled(): return {}

        # --- LANGKAH 1: REKLASIFIKASI INDEKS BENCANA ---
        feedback.pushInfo("Langkah 1: Merekalasifikasi Indeks Bencana...")
        kelas_bencana_path = os.path.join(out_folder, "Kelas_Bencana.tif")
        # FIX #1: Tambahkan filter (A>0) untuk mengecualikan piksel NoData/background.
        # Formula lama: piksel bernilai 0 akan salah diklasifikasikan sebagai Kelas 1 (Rendah)
        # karena kondisi (A<=0.333) terpenuhi. Formula baru hanya memproses nilai > 0.
        calc_expression = "(A>0)*(A<=0.333)*1 + (A>0.333)*(A<=0.666)*2 + (A>0.666)*3"
        # FIX #6: native:rastercalculator tidak tersedia di QGIS 3.34+.
        # Ganti dengan gdal:rastercalculator yang stabil di semua versi QGIS 3.x
        calc_params = {
            'INPUT_A': parameters[self.P_RASTER_BENCANA],
            'BAND_A': 1,
            'FORMULA': calc_expression,
            'NO_DATA': 0,
            'RTYPE': 5,  # 5 = Float32
            'OUTPUT': kelas_bencana_path
        }
        processing.run("gdal:rastercalculator", calc_params, context=context, feedback=feedback, is_child_algorithm=True)
        feedback.setProgress(30)

        if feedback.isCanceled(): return {}

        # --- LANGKAH 2: ZONAL HISTOGRAM (Luasan Bahaya per Desa) ---
        feedback.pushInfo("Langkah 2: Menghitung Luas Kelas Bencana per Desa...")
        id_desa_path = os.path.join(out_folder, "ID_Desa.tif")
        rasterize_params = {
            'INPUT': aligned_vektor,
            'FIELD': field_id,
            'UNITS': 1,
            'WIDTH': pixel_size_x,
            'HEIGHT': pixel_size_y,
            'EXTENT': f"{extent_bencana.xMinimum()},{extent_bencana.xMaximum()},{extent_bencana.yMinimum()},{extent_bencana.yMaximum()}",
            'NODATA': -9999,
            'OUTPUT': id_desa_path
        }
        processing.run("gdal:rasterize", rasterize_params, context=context, feedback=feedback, is_child_algorithm=True)
        
        luas_per_pixel_ha = (pixel_size_x * pixel_size_y) / 10000.0
        
        hist_params = {
            'INPUT_RASTER': kelas_bencana_path,
            'RASTER_BAND': 1,
            'INPUT_VECTOR': aligned_vektor,
            'COLUMN_PREFIX': 'KL_',
            'OUTPUT': 'memory:'
        }
        res_hist = processing.run("native:zonalhistogram", hist_params, context=context, feedback=feedback, is_child_algorithm=True)
        # FIX #7: is_child_algorithm=True mengembalikan string ID, bukan layer object.
        # Gunakan context.getMapLayer() untuk mendapatkan layer yang sebenarnya.
        hist_layer = context.getMapLayer(res_hist['OUTPUT'])
        if hist_layer is None:
            raise QgsProcessingException("Gagal memuat layer hasil Zonal Histogram dari context.")
        feedback.setProgress(60)

        if feedback.isCanceled(): return {}

        # --- LANGKAH 3: ZONAL STATISTICS PENDUDUK ---
        feedback.pushInfo("Langkah 3: Menghitung Penduduk Terpapar...")
        zona_gabungan_path = os.path.join(out_folder, "Zona_Gabungan.tif")
        # CATATAN #5 (Validasi ID Desa): Formula encoding ini (ID_Desa*10 + Kelas) aman
        # selama nilai ID Desa tidak terlalu besar. Raster GeoTIFF Float32 dapat menampung
        # integer hingga ~16 juta (2^24) dengan presisi penuh. Jika menggunakan Kode BPS
        # 10 digit (misal 3215010001), nilai zona akan melebihi batas Float32 dan menyebabkan
        # kehilangan presisi. Gunakan field ID sekuensial (integer kecil) jika ini terjadi.
        calc_zona_exp = "(A * 10) + B"
        calc_zona_params = {
            'INPUT_A': id_desa_path, 'BAND_A': 1,
            'INPUT_B': kelas_bencana_path, 'BAND_B': 1,
            'FORMULA': calc_zona_exp,
            'NO_DATA': 0,
            'RTYPE': 5,  # 5 = Float32
            'OUTPUT': zona_gabungan_path
        }
        processing.run("gdal:rastercalculator", calc_zona_params, context=context, feedback=feedback, is_child_algorithm=True)

        polygon_zona_path = os.path.join(out_folder, "Poligon_Zona.shp")
        poly_params = {
            'INPUT': zona_gabungan_path,
            'BAND': 1,
            'FIELD': 'ZONA_VAL',
            'EIGHT_CONNECTEDNESS': False,
            'OUTPUT': polygon_zona_path
        }
        res_poly = processing.run("gdal:polygonize", poly_params, context=context, feedback=feedback, is_child_algorithm=True)
        
        stat_params = {
            'INPUT': res_poly['OUTPUT'],
            'INPUT_RASTER': aligned_penduduk,
            'RASTER_BAND': 1,
            'COLUMN_PREFIX': 'POP_',
            'STATISTICS': [1], 
            'OUTPUT': 'memory:'
        }
        res_stat = processing.run("native:zonalstatisticsfb", stat_params, context=context, feedback=feedback, is_child_algorithm=True)
        # FIX #7 (sama): resolve string ID ke layer object
        stat_layer = context.getMapLayer(res_stat['OUTPUT'])
        if stat_layer is None:
            raise QgsProcessingException("Gagal memuat layer hasil Zonal Statistics dari context.")
        
        feedback.setProgress(80)

        # Parsing hasil Zonal Stat
        penduduk_dict = {}
        for feat in stat_layer.getFeatures():
            zona_val = feat['ZONA_VAL']
            if zona_val is None or zona_val < 0: continue
            
            kelas_bencana = int(zona_val % 10)
            id_desa = int(zona_val // 10)
            pop_sum = feat['POP_sum'] or 0
            
            if id_desa not in penduduk_dict:
                penduduk_dict[id_desa] = {1: 0, 2: 0, 3: 0}
            if kelas_bencana in [1, 2, 3]:
                penduduk_dict[id_desa][kelas_bencana] += pop_sum

        # --- PERSIAPAN TABEL FINAL ---
        feedback.pushInfo("Menyusun data untuk laporan...")
        
        # FIX #3: Ekstrak daftar nama field sekali di luar loop untuk efisiensi.
        # Sebelumnya, [f.name() for f in hist_layer.fields()] dipanggil 3x per fitur,
        # menyebabkan overhead O(n*m) yang tidak perlu.
        hist_field_names = [f.name() for f in hist_layer.fields()]
        
        data_rows = []
        for feat in hist_layer.getFeatures():
            f_id = feat[field_id]
            if f_id is None: continue
            
            try: id_val = int(f_id)
            except ValueError: id_val = str(f_id)
            
            nm_desa = str(feat[field_desa]) if feat[field_desa] else "-"
            nm_kec = str(feat[field_kec]) if feat[field_kec] else "-"
            
            # Gunakan daftar field yang sudah diekstrak di luar loop
            l1_px = feat['KL_1'] if 'KL_1' in hist_field_names else 0
            l2_px = feat['KL_2'] if 'KL_2' in hist_field_names else 0
            l3_px = feat['KL_3'] if 'KL_3' in hist_field_names else 0
            
            l1 = round((l1_px or 0) * luas_per_pixel_ha, desimal_luas)
            l2 = round((l2_px or 0) * luas_per_pixel_ha, desimal_luas)
            l3 = round((l3_px or 0) * luas_per_pixel_ha, desimal_luas)
            
            p1, p2, p3 = 0, 0, 0
            key_id = id_val if id_val in penduduk_dict else (int(id_val) if isinstance(id_val, str) and id_val.isdigit() and int(id_val) in penduduk_dict else None)
            if key_id is not None:
                p1 = round(penduduk_dict[key_id].get(1, 0))
                p2 = round(penduduk_dict[key_id].get(2, 0))
                p3 = round(penduduk_dict[key_id].get(3, 0))
            
            # FIX #8: Desa tanpa area bahaya (l1=l2=l3=0) dilewati sepenuhnya —
            # tidak muncul di laporan Excel maupun HTML.
            # Sekaligus mencegah penduduk "hantu" dari konflik piksel perbatasan.
            if l1 == 0 and l2 == 0 and l3 == 0:
                continue
            
            data_rows.append({
                'kec': nm_kec, 'desa': nm_desa,
                
                'l_1': l1, 'l_2': l2, 'l_3': l3,
                'p_1': p1, 'p_2': p2, 'p_3': p3
            })

        data_rows.sort(key=lambda x: (x['kec'], x['desa']))

        # --- EKSPOR OUTPUT ---
        self.export_to_excel(data_rows, out_excel, feedback)
        
        result_dict = {self.P_OUT_EXCEL: out_excel}

        if out_html:
            self.export_to_html(data_rows, out_html, feedback)
            result_dict[self.P_OUT_HTML] = out_html

        feedback.setProgress(100)
        return result_dict

    def export_to_excel(self, data_rows, filepath, feedback):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dampak Bencana"

        dark_blue = PatternFill("solid", fgColor="002060")
        light_blue = PatternFill("solid", fgColor="95B3D7")
        dark_green = PatternFill("solid", fgColor="275720")
        green_fill = PatternFill("solid", fgColor="92D050")
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        red_fill = PatternFill("solid", fgColor="FF0000")

        white_bold = Font(color="FFFFFF", bold=True)
        black_bold = Font(color="000000", bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        double_top = Border(left=thin, right=thin, top=Side(border_style="double"), bottom=thin)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = {
            'A1:A4': ('DESA/KELURAHAN', dark_blue, white_bold),
            'B1:B4': ('KECAMATAN', dark_blue, white_bold),
            'C1:G1': ('BAHAYA', light_blue, black_bold),
            'H1:K1': ('Penduduk Terpapar', dark_green, white_bold),
            'C2:E3': ('LUAS BAHAYA (HA)', light_blue, black_bold),
            'F2:F4': ('TOTAL LUAS', light_blue, black_bold),
            'G2:G4': ('KELAS', light_blue, black_bold),
            'H2:J2': ('POTENSI PENDUDUK TERPAPAR', dark_green, white_bold),
            'H3:J3': ('JUMLAH PENDUDUK TERPAPAR PER KELAS BAHAYA', dark_green, white_bold),
            'K2:K4': ('TOTAL JUMLAH PENDUDUK TERPAPAR (JIWA)', dark_green, white_bold),
            'C4': ('RENDAH', green_fill, black_bold),
            'D4': ('SEDANG', yellow_fill, black_bold),
            'E4': ('TINGGI', red_fill, white_bold),
            'H4': ('RENDAH', green_fill, black_bold),
            'I4': ('SEDANG', yellow_fill, black_bold),
            'J4': ('TINGGI', red_fill, white_bold),
        }

        for cell_range, (text, fill, font) in headers.items():
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            if ':' in cell_range:
                ws.merge_cells(cell_range)

        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=11):
            for cell in row:
                cell.alignment = center_align
                cell.border = border
                for cell_range, (text, fill, font) in headers.items():
                    if ':' in cell_range:
                        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(cell_range)
                        if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                            cell.fill = fill; cell.font = font
                    else:
                        if cell.coordinate == cell_range:
                            cell.fill = fill; cell.font = font

        start_row = 5
        for i, row_data in enumerate(data_rows):
            cur_row = start_row + i
            
            ws.cell(row=cur_row, column=1, value=row_data['desa']).border = border
            ws.cell(row=cur_row, column=2, value=row_data['kec']).border = border
            ws.cell(row=cur_row, column=3, value=row_data['l_1']).border = border
            ws.cell(row=cur_row, column=4, value=row_data['l_2']).border = border
            ws.cell(row=cur_row, column=5, value=row_data['l_3']).border = border
            
            f_cell = ws.cell(row=cur_row, column=6, value=f"=SUM(C{cur_row}:E{cur_row})")
            f_cell.border = border
            
            # FIX #4: Jika dua atau lebih kelas memiliki luas sama (tie), formula
            # memprioritaskan: Rendah > Sedang > Tinggi (urutan pengecekan IF).
            # Ini adalah perilaku yang disengaja: sisi konservatif (prioritas kelas lebih rendah).
            g_cell = ws.cell(row=cur_row, column=7, value=f'=IF(MAX(C{cur_row}:E{cur_row})=C{cur_row},"Rendah",IF(MAX(C{cur_row}:E{cur_row})=D{cur_row},"Sedang","Tinggi"))')
            g_cell.border = border
            
            ws.cell(row=cur_row, column=8, value=row_data['p_1']).border = border
            ws.cell(row=cur_row, column=9, value=row_data['p_2']).border = border
            ws.cell(row=cur_row, column=10, value=row_data['p_3']).border = border
            
            k_cell = ws.cell(row=cur_row, column=11, value=f"=SUM(H{cur_row}:J{cur_row})")
            k_cell.border = border

        total_row = start_row + len(data_rows)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        total_label = ws.cell(row=total_row, column=1, value="TOTAL KESELURUHAN")
        total_label.font = black_bold; total_label.alignment = center_align
        ws.cell(row=total_row, column=2).border = double_top
        total_label.border = double_top

        cols_to_sum = ['C', 'D', 'E', 'F', 'H', 'I', 'J', 'K']
        for col_letter in cols_to_sum:
            cell = ws[f"{col_letter}{total_row}"]
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{total_row-1})"
            cell.font = black_bold; cell.border = double_top
            
        ws[f"G{total_row}"].border = double_top
        
        widths = {'A': 22, 'B': 18, 'C': 12, 'D': 12, 'E': 12, 'F': 15, 'G': 15, 'H': 12, 'I': 12, 'J': 12, 'K': 18}
        for col, w in widths.items(): ws.column_dimensions[col].width = w
            
        ws.freeze_panes = "A5"

        try:
            wb.save(filepath)
        except PermissionError:
            raise QgsProcessingException(f"PermissionError: Tidak dapat menyimpan file excel.")

    def export_to_html(self, data_rows, filepath, feedback):
        feedback.pushInfo("Mengekspor laporan ke HTML Interaktif...")
        
        html_template = """<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Laporan Analisis Dampak Bencana</title>
    <!-- Bootstrap CSS -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <!-- DataTables CSS -->
    <link href="https://cdn.datatables.net/1.13.6/css/dataTables.bootstrap5.min.css" rel="stylesheet">
    <style>
        body { background-color: #f8f9fa; padding: 20px; font-size: 0.9rem; }
        .card { box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: none; }
        .table thead th { background-color: #002060; color: white; vertical-align: middle; text-align: center; }
        .bg-rendah { background-color: #d4edda !important; color: #155724; }
        .bg-sedang { background-color: #fff3cd !important; color: #856404; }
        .bg-tinggi { background-color: #f8d7da !important; color: #721c24; }
    </style>
</head>
<body>
    <div class="container-fluid">
        <h2 class="mb-4 text-center">Laporan Analisis Dampak Bencana Dasimetrik</h2>
        <div class="card p-4">
            <div class="table-responsive">
                <table id="tabelBencana" class="table table-bordered table-hover w-100">
                    <thead>
                        <tr>
                            <th rowspan="2">Desa/Kelurahan</th>
                            <th rowspan="2">Kecamatan</th>
                            <th colspan="3" class="bg-primary">Luas Bahaya (Ha)</th>
                            <th rowspan="2">Total Luas</th>
                            <th rowspan="2">Kelas Dominan</th>
                            <th colspan="3" class="bg-success">Penduduk Terpapar (Jiwa)</th>
                            <th rowspan="2">Total Penduduk</th>
                        </tr>
                        <tr>
                            <th style="background-color: #92D050; color: black;">Rendah</th>
                            <th style="background-color: #FFFF00; color: black;">Sedang</th>
                            <th style="background-color: #FF0000; color: white;">Tinggi</th>
                            <th style="background-color: #92D050; color: black;">Rendah</th>
                            <th style="background-color: #FFFF00; color: black;">Sedang</th>
                            <th style="background-color: #FF0000; color: white;">Tinggi</th>
                        </tr>
                    </thead>
                    <tbody>
"""
        
        # Populate Table Rows
        for row in data_rows:
            tot_luas = round(row['l_1'] + row['l_2'] + row['l_3'], 2)
            tot_pop = row['p_1'] + row['p_2'] + row['p_3']
            
            # Cari Kelas Dominan
            max_luas = max(row['l_1'], row['l_2'], row['l_3'])
            if max_luas == 0:
                kelas_dom = "-"
                dom_class = ""
            elif max_luas == row['l_1']:
                kelas_dom = "Rendah"
                dom_class = "bg-rendah"
            elif max_luas == row['l_2']:
                kelas_dom = "Sedang"
                dom_class = "bg-sedang"
            else:
                kelas_dom = "Tinggi"
                dom_class = "bg-tinggi"

            html_template += f"""
                        <tr>
                            <td>{row['desa']}</td>
                            <td>{row['kec']}</td>
                            <td class="text-end">{row['l_1']:,}</td>
                            <td class="text-end">{row['l_2']:,}</td>
                            <td class="text-end">{row['l_3']:,}</td>
                            <td class="text-end fw-bold">{tot_luas:,}</td>
                            <td class="text-center {dom_class} fw-bold">{kelas_dom}</td>
                            <td class="text-end">{row['p_1']:,}</td>
                            <td class="text-end">{row['p_2']:,}</td>
                            <td class="text-end">{row['p_3']:,}</td>
                            <td class="text-end fw-bold">{tot_pop:,}</td>
                        </tr>"""

        # Close Table and Add Scripts
        html_template += """
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <!-- jQuery & DataTables JS -->
    <script src="https://code.jquery.com/jquery-3.7.0.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/dataTables.bootstrap5.min.js"></script>
    <script>
        $(document).ready(function() {
            $('#tabelBencana').DataTable({
                "language": {
                    "url": "//cdn.datatables.net/plug-ins/1.13.6/i18n/id.json"
                },
                "pageLength": 25,
                "order": [[ 1, "asc" ], [ 0, "asc" ]] // Sort by Kecamatan then Desa
            });
        });
    </script>
</body>
</html>
"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_template)